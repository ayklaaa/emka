from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import user_passes_test
from django.contrib import messages
from django.contrib.auth import get_user_model
from orders.models import Order
from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from openpyxl.chart import BarChart, Reference
from collections import defaultdict
from decimal import Decimal
import io
from datetime import datetime

User = get_user_model()


def is_admin(user):
    """Проверка, является ли пользователь администратором"""
    return user.is_authenticated and user.is_staff


@user_passes_test(is_admin)
def admin_dashboard(request):
    """Панель администратора"""
    return render(request, 'custom_admin/dashboard.html')


@user_passes_test(is_admin)
def order_management(request):
    """Управление заказами"""
    orders = Order.objects.select_related('user').all().order_by('-created_at')

    if request.method == 'POST':
        try:
            order_id = request.POST.get('order_id')
            new_status = request.POST.get('status')
            order = get_object_or_404(Order, id=order_id)
            order.status = new_status
            order.save()
            messages.success(request, 'Статус заказа успешно обновлен')
        except Exception as e:
            messages.error(request, f'Ошибка при обновлении заказа: {str(e)}')

        return redirect('custom_admin:order_management')

    return render(request, 'custom_admin/order_management.html', {'orders': orders})


@user_passes_test(is_admin)
def user_management(request):
    """Управление пользователями"""
    users = User.objects.all().order_by('-date_joined')

    if request.method == 'POST':
        try:
            user_id = request.POST.get('user_id')
            action = request.POST.get('action')
            user = get_object_or_404(User, id=user_id)

            if action == 'toggle_staff':
                user.is_staff = not user.is_staff
                user.save()
                msg = 'Статус персонала изменен'
            elif action == 'toggle_active':
                user.is_active = not user.is_active
                user.save()
                msg = 'Статус активности изменен'

            messages.success(request, msg)
        except Exception as e:
            messages.error(request, f'Ошибка при обновлении пользователя: {str(e)}')

        return redirect('custom_admin:user_management')

    return render(request, 'custom_admin/user_management.html', {'users': users})


@user_passes_test(is_admin)
def generate_sales_report(request):
    """Генерация отчета о продажах в Excel с диаграммой"""
    # Создаем Excel-файл в памяти
    output = io.BytesIO()
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Отчет о продажах"

    # Настройка стилей
    header_font = Font(bold=True)
    center_alignment = Alignment(horizontal='center')

    # Заголовки столбцов
    headers = ['ID заказа', 'Пользователь', 'Статус', 'Дата создания', 'Сумма (руб)']
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        cell = worksheet[f'{col_letter}1']
        cell.value = header
        cell.font = header_font
        cell.alignment = center_alignment
        worksheet.column_dimensions[col_letter].width = 20

    # Получаем и обрабатываем заказы
    orders = Order.objects.select_related('user').all().order_by('-created_at')
    status_totals = defaultdict(Decimal)
    monthly_totals = defaultdict(Decimal)

    for row_num, order in enumerate(orders, 2):
        worksheet.cell(row=row_num, column=1, value=order.id).alignment = center_alignment
        worksheet.cell(row=row_num, column=2, value=str(order.user))
        worksheet.cell(row=row_num, column=3, value=order.get_status_display()).alignment = center_alignment
        worksheet.cell(row=row_num, column=4,
                       value=order.created_at.strftime('%Y-%m-%d %H:%M')).alignment = center_alignment
        worksheet.cell(row=row_num, column=5, value=float(order.total_price)).alignment = center_alignment

        # Статистика по статусам и месяцам
        status_totals[order.get_status_display()] += order.total_price
        month_key = order.created_at.strftime('%Y-%m')
        monthly_totals[month_key] += order.total_price

    # Добавляем раздел статистики по статусам
    start_row = len(orders) + 4
    worksheet.cell(row=start_row - 1, column=1, value="Статистика по статусам").font = header_font

    worksheet.cell(row=start_row, column=1, value="Статус").font = header_font
    worksheet.cell(row=start_row, column=2, value="Сумма (руб)").font = header_font

    for i, (status, total) in enumerate(status_totals.items(), start=start_row + 1):
        worksheet.cell(row=i, column=1, value=status)
        worksheet.cell(row=i, column=2, value=float(total))

    # Диаграмма по статусам
    chart1 = BarChart()
    chart1.title = "Продажи по статусу заказа"
    chart1.x_axis.title = "Статус"
    chart1.y_axis.title = "Сумма (руб)"

    data = Reference(worksheet, min_col=2, min_row=start_row, max_row=start_row + len(status_totals))
    cats = Reference(worksheet, min_col=1, min_row=start_row + 1, max_row=start_row + len(status_totals))
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    worksheet.add_chart(chart1, "E2")

    # Добавляем раздел статистики по месяцам
    monthly_start_row = start_row + len(status_totals) + 3
    worksheet.cell(row=monthly_start_row - 1, column=1, value="Статистика по месяцам").font = header_font

    worksheet.cell(row=monthly_start_row, column=1, value="Месяц").font = header_font
    worksheet.cell(row=monthly_start_row, column=2, value="Сумма (руб)").font = header_font

    for i, (month, total) in enumerate(sorted(monthly_totals.items()), start=monthly_start_row + 1):
        worksheet.cell(row=i, column=1, value=month)
        worksheet.cell(row=i, column=2, value=float(total))

    # Диаграмма по месяцам
    chart2 = BarChart()
    chart2.title = "Продажи по месяцам"
    chart2.x_axis.title = "Месяц"
    chart2.y_axis.title = "Сумма (руб)"

    data = Reference(worksheet, min_col=2, min_row=monthly_start_row, max_row=monthly_start_row + len(monthly_totals))
    cats = Reference(worksheet, min_col=1, min_row=monthly_start_row + 1,
                     max_row=monthly_start_row + len(monthly_totals))
    chart2.add_data(data, titles_from_data=True)
    chart2.set_categories(cats)
    worksheet.add_chart(chart2, "E20")

    # Сохраняем и возвращаем файл
    workbook.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"sales_report_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    return response


def order_change(request, order_id):
    """Функция для изменения заказа в админке"""
    order = get_object_or_404(Order, id=order_id)

    if request.method == 'POST':
        # Обработка изменения заказа
        status = request.POST.get('status')
        if status:
            order.status = status
            order.save()
            messages.success(request, 'Заказ успешно обновлен')
            return redirect('custom_admin:order_management')

    context = {
        'order': order,
        'status_choices': Order.STATUS_CHOICES
    }
    return render(request, 'custom_admin/order_change.html', context)